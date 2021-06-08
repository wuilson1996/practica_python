from tkinter import *
import openpyxl
import os
import os.path as path
from datetime import datetime
from tkinter import messagebox as MessageBox
from tkinter import messagebox, filedialog


#se inicia la clase facturacion
class facturacion():
    #se inicia la ventana
    def __init__(self,tam):


        lista=["Cantidad","Descripcion","Valor Unit.","Valor Total"]
        lista2=["SUBTOTAL","ABONO","TOTAL ->"]
        self.ventana=Tk()
        self.ventana.iconbitmap("icono.ico")
        self.ventana.title("Sistema de Facturacion Version 1.0.0")
        self.ventana.geometry("730x650")
        self.ventana.resizable(0,0)
        self.color="#ff9983"
        self.color2="white" 
        self.color3="#fccd9f"
        self.ventana.configure(bg=self.color2)
        self.tam=tam
        self.file=''
        now=datetime.now()
        fecha=now.day ,'/', now.month ,'/', now.year
        
        self.Clien=StringVar()
        self.VarSede=StringVar()
        self.VarNit=StringVar()
        self.VarTlf=StringVar()
        self.VarCel=StringVar()
        self.Var_Direc=StringVar()

        self.F_Venta=Label(self.ventana,text="Cliente",bg=self.color2)
        self.F_Venta.place(x=490,y=20)
        self.EClient=Entry(self.ventana, textvariable=self.Clien,justify=CENTER)
        self.EClient.place(x=550, y=20, width=135)

        self.Sede=Label(self.ventana,text="Sede",bg=self.color2)
        self.Sede.place(x=490,y=40)
        self.ESede=Entry(self.ventana, textvariable=self.VarSede,justify=CENTER)
        self.ESede.place(x=550, y=40, width=135)

        self.Fecha_Actual=Label(self.ventana,text="CC/Nit",bg=self.color2)
        self.Fecha_Actual.place(x=490,y=60)
        self.ECel=Entry(self.ventana, textvariable=self.VarNit,justify=CENTER)
        self.ECel.place(x=550, y=60, width=135)

        self.Fecha_Entrega=Label(self.ventana,text="Telefono",bg=self.color2)
        self.Fecha_Entrega.place(x=490,y=80)
        self.ECel=Entry(self.ventana, textvariable=self.VarTlf,justify=CENTER)
        self.ECel.place(x=550, y=80, width=135)

        #seccion de nit y tlf
        self.Cel=Label(self.ventana, text="Celular",bg=self.color2)
        self.Cel.place(x=490,y=100)
        self.ECel=Entry(self.ventana, textvariable=self.VarCel,justify=CENTER)
        self.ECel.place(x=550, y=100, width=135)
        
        self.Direc=Label(self.ventana, text="Direccion",bg=self.color2)
        self.Direc.place(x=490,y=120)
        self.EDirec=Entry(self.ventana, textvariable=self.Var_Direc,justify=CENTER)
        self.EDirec.place(x=550,y=120, width=135)

        self.Num_Fact=StringVar()
        self.VarFecha=StringVar()
        self.Fecha_A=StringVar()
        self.VarTipo=StringVar()
        self.Fecha_A.set(fecha)

        self.Telf=Label(self.ventana,text="Telefono:       310 735 45 16",bg=self.color2)
        self.Telf.place(x=240,y=20)

        self.Wsp=Label(self.ventana,text="Whatsapp:     301 210 19 41",bg=self.color2)
        self.Wsp.place(x=240,y=40)

        self.N_Factura=Label(self.ventana,text="N° Factura:",bg=self.color2)
        self.N_Factura.place(x=240,y=60)
        self.ECel=Entry(self.ventana, textvariable=self.Num_Fact,justify=CENTER)
        self.ECel.place(x=300, y=60, width=135)

        self.Fecha_Actual=Label(self.ventana,text="Fecha:",bg=self.color2)
        self.Fecha_Actual.place(x=240,y=80)
        self.ECel=Entry(self.ventana, textvariable=self.Fecha_A,justify=CENTER)
        self.ECel.place(x=300, y=80, width=135)

        #seccion de nit y tlf
        self.Cel=Label(self.ventana, text="Tipo",bg=self.color2)
        self.Cel.place(x=240,y=100)
        self.ECel=Entry(self.ventana, textvariable=self.VarTipo,justify=CENTER)
        self.ECel.place(x=300, y=100, width=135)
        
        self.Direc=Label(self.ventana, text="Jefe:                Juan Jose Echavarria",bg=self.color2)
        self.Direc.place(x=240,y=120)
        

        self.frame2=Frame(self.ventana)
        self.frame2.place(x=50,y=150)

        self.frame=Frame(self.frame2)
        self.frame.grid(row=0, column=0)
        self.frame.configure(bg=self.color)

        self.frame3=Frame(self.frame2)
        self.frame3.grid(row=1,column=0,pady=5, sticky="w")
        self.frame3.configure(bg=self.color)

        self.frame4=Frame(self.frame2)
        self.frame4.grid(row=1,column=0,sticky="e")

        self.logo=PhotoImage(file="logo.png")
        self.Logo=Label(self.ventana,image=self.logo, bg=self.color2)
        self.Logo.place(x=50, y=10)

        self.Direc=Label(self.ventana, text="Direccion: Calle 57F N° 92 DD-03",bg=self.color2)
        self.Direc.place(x=45,y=120)
        
       

        for i in range(4):

            self.texto=Label(self.frame,text=lista[i], bg=self.color)
            self.texto.grid(row=0,column=i, ipady=5)

        self.llenar()

        self.pie=Entry(self.frame3)
        self.pie.grid(row=0, column=0, ipadx=35, ipady=30, rowspan=3)
        
        self.pie2=Entry(self.frame3)
        self.pie2.grid(row=0, column=1, ipadx=35, ipady=30,rowspan=3)
        
        for i in range(3):
            if(i!=1):
                self.Ttotal=Label(self.frame3,text=lista2[i],bg=self.color)
                self.Ttotal.grid(row=i,column=2, ipadx=30)

            else:
                self.Ttotal=Label(self.frame3,text=lista2[i],bg=self.color2)
                self.Ttotal.grid(row=i,column=2, ipadx=30)

        self.BarraMenu()

        self.ventana.mainloop()


    def BarraMenu(self):

        barra_Menu=Menu(self.ventana)

        self.ventana.config(menu=barra_Menu, width=300, height=300)
        
        archivoMenu=Menu(barra_Menu, tearoff=0)
        archivoMenu.add_command(label="Nuevo", command=lambda:Nuevo(self.ventana, self.aux))
        archivoMenu.add_command(label="Guardar", command=self.Sobre_Escribir)
        archivoMenu.add_command(label="Guardar como", command=self.Guar_Como)
        archivoMenu.add_separator()
        archivoMenu.add_command(label="Cerrar", command=self.cerrar_Documento)
        archivoMenu.add_command(label="Salir", command=self.salir_Aplicacion)

        archivoEdicion=Menu(barra_Menu, tearoff=0)
        archivoEdicion.add_command(label="modificar factura", command=self.cargar_informacion)
        archivoEdicion.add_command(label="Factura", command=lambda: self.Tipo(1))
        archivoEdicion.add_command(label="Cotizacion", command=lambda: self.Tipo(2))
        #archivoEdicion.add_command(label="Pegar")
        #archivoEdicion.add_command(label="Cortar")

        archivoHerramientas=Menu(barra_Menu, tearoff=0)
        archivoHerramientas.add_command(label="Numero-Factura",command=self.cargar_Num_fac)
        archivoHerramientas.add_command(label="Crear-Archivo",command=self.Crear_Archivo)
        archivoHerramientas.add_command(label="Tamaño de Factura",command=lambda:Cambiar_Tam(self.ventana,self.aux))
        
        Colores=Menu(archivoHerramientas, tearoff=0)
        Colores.add_command(label="Normal", command=lambda: self.Cambiar_color(0))
        Colores.add_command(label="Color1", command=lambda: self.Cambiar_color(1))
        Colores.add_command(label="Color2", command=lambda: self.Cambiar_color(2))
        Colores.add_command(label="Color3", command=lambda: self.Cambiar_color(3))
        Colores.add_command(label="Color4", command=lambda: self.Cambiar_color(4))
        Colores.add_command(label="Color5", command=lambda: self.Cambiar_color(5))
        Colores.add_command(label="Color6", command=lambda: self.Cambiar_color(6))
        Colores.add_command(label="Color7", command=lambda: self.Cambiar_color(7))
        Colores.add_command(label="Color8", command=lambda: self.Cambiar_color(8))
        
        archivoAyuda=Menu(barra_Menu, tearoff=0)
        archivoAyuda.add_command(label="Licencia", command=self.aviso_Licencia)
        archivoAyuda.add_command(label="Documentacion", command=self.Documentacion)
        archivoAyuda.add_command(label="Acerca de...", command=self.info_Adicional)

        barra_Menu.add_cascade(label="Archivo", menu=archivoMenu)

        barra_Menu.add_cascade(label="Edicion", menu=archivoEdicion)

        barra_Menu.add_cascade(label="Herramientas", menu=archivoHerramientas)

        barra_Menu.add_cascade(label="Ayuda", menu=archivoAyuda)

        archivoHerramientas.add_cascade(label="Colores",menu=Colores)

    def Documentacion(self):

        hoja=Tk()
        hoja.geometry("600x500")
        hoja.iconbitmap("icono.ico")
        hoja.title("Documentacion")
        Label(hoja,text="configuracion:").place(x=10,y=50)
        Label(hoja,text="1: cree una carpeta en documento llamada-facturacion").place(x=10,y=70)
        Label(hoja,text="2: Copie el contenido de el programa en la carpeta creada").place(x=10,y=90)
        Label(hoja,text="3: cree una carpeta llamada facturas, el cual es donde se guardaran sus facturas").place(x=10,y=110)
        Label(hoja,text="4: click derecho en el archivo exe que brevemente copio en facturacion/enviar/acceso directo a escritorio").place(x=10,y=130)
        Label(hoja,text="5: ir a la opcion herramientas en el menu/crear-archivo").place(x=10,y=150)

        hoja.mainloop()

    def Tipo(self,aux):

        if aux == 1:
            self.VarTipo.set("Factura")
        else:
            self.VarTipo.set("Cotizacion")

    def Cambiar_color(self,i):
        color=["white","#FEED71","#C2FE71","#71FEF3","#71A2FE","#9171FE","#B871FE","#FE9171","#FE87DA"]

        self.color2=color[i]
        self.ventana.configure(bg=self.color2)
        self.tlf.configure(bg=self.color2)
        self.Nit.configure(bg=self.color2)
        self.direccion.configure(bg=self.color2)
        self.cliente.configure(bg=self.color2)
        self.F_Venta.configure(bg=self.color2)
        self.Coti.configure(bg=self.color2)
        self.N_FACT.configure(bg=self.color2)
        self.Fecha_Actual.configure(bg=self.color2)
        self.Fecha_Entrega.configure(bg=self.color2)
        self.Logo.configure(bg=self.color2)


    def Sobre_Escribir(self):

        if(self.file!=''):

            result=MessageBox.askokcancel("Guardar Factura","¿Desea Guardar la Factura?")
           
            if result == True:
                lis=["A","B","G","H"]
                excel_doc=openpyxl.load_workbook(self.file)
                hoja=excel_doc.get_sheet_by_name('Hoja1')

                for i in range(self.tam):  #ciclo para iniciar varibles para almacenar 
            
                    for j in range(4):
                       
                        try:

                            if self.info_doc[i][0].get() > 0:

                                hoja[lis[j]+str(i+9)]=self.info_doc[i][j].get()

                            else:

                                hoja[lis[1]+str(i+9)]=self.info_doc[i][1].get()
                                j=4

                        except:

                            hoja[lis[1]+str(i+9)]=self.info_doc[i][1].get()
                            j=4

                    #print("\n")
                fecha=[]
                hoja["D3"]=self.Num_Fact.get()
            
                for i in self.Fecha_A.get():
                
                    if(i != "," and i != "(" and i != ")" and i != "'" and i !=" "):
                        fecha.append(i)
            
            
                hoja["D4"]=''.join(fecha)
                hoja["D6"]=self.VarTipo.get()

                for i in range(3):

                    hoja[lis[3]+str(i+23)]=self.VarTotal[i].get()

                hoja["G1"]=self.Clien.get()
                hoja["G2"]=self.VarSede.get()
                hoja["G3"]=self.VarNit.get()
                hoja["G4"]=self.VarTlf.get()
                hoja["G5"]=self.VarCel.get()
                hoja["G6"]=self.Var_Direc.get()

                excel_doc.save(self.file)

                messagebox.showinfo("Guardar Factura","Se ha guardado con exito la Factura")
        else:

            self.Guar_Como()
            


    def cerrar_Documento(self):
        valor=messagebox.askretrycancel("Reintentar","No es posible cerrar. Documento bloqueado")

    def salir_Aplicacion(self):
        valor=messagebox.askokcancel("Salir","¿Desea salir de la aplicacion?")
        
        if valor==True:
            valor=messagebox.showinfo("Gracias","Gracias Por usar nuestro Software")
            
            if(valor=='ok'):
                self.ventana.destroy()

    def aviso_Licencia(self):
        messagebox.showwarning("Licencia", "producto bajo licencia GNU")

    def info_Adicional(self):
        messagebox.showinfo("Desarrollado por wuil-Soft","Sistema de Facturacion Version 1.0.0 / Fecha: 09/01/2020")
    
    #se encarga de llenar la cantidad de filas
    def llenar(self):

        self.info_doc=[]
        self.VarTotal=[]

        for i in range(3):
            self.VarTotal.append(DoubleVar())

        for i in range(self.tam):  #ciclo para iniciar varibles para almacenar 
            self.info_doc.append([])
			
            for j in range(4):
                
                if(j==1):           
                    
                    self.info_doc[i].append(StringVar())
                else:

                    self.info_doc[i].append(DoubleVar())
        
        #llamado de la funcion para ir actualizando el valor total       
        self.Actualizar()
            

        for i in range(self.tam):  #ciclo para iniciar 
			
            for j in range(4):
               
                if(j==1):
                    entry=Entry(self.frame, textvariable=self.info_doc[i][j])
                    entry.grid(row=i+1, column=j, ipadx=70, ipady=5)
                
                else:
                    entry=Entry(self.frame, textvariable=self.info_doc[i][j], justify=RIGHT)
                    entry.grid(row=i+1, column=j, ipady=5)

        for i in range(3):

            self.total=Entry(self.frame4,textvariable=self.VarTotal[i])
            self.total.grid(row=i,column=0, ipady=4)
                

    def Guar_Como(self):
        
        nombre_arch=''
        nombre_arch=filedialog.asksaveasfilename()
    
        self.Guardar_como(nombre_arch)
   
    #funcion: se encarga de guardar los datos ingresados
    def Guardar_como(self, nombre_arch):
        
        if(nombre_arch!=''):
            self.nombre_arch=nombre_arch

            result=MessageBox.askokcancel("Guardar Factura","¿Desea Guardar la Factura?")
            #print(self.nombre_arch)
            if result == True:
                lis=["A","B","G","H"]
                excel_doc=openpyxl.load_workbook('juan.xlsx')
                hoja=excel_doc.get_sheet_by_name('Hoja1')

                for i in range(self.tam):  #ciclo para iniciar varibles para almacenar 
			
                    for j in range(4):
                    
                        try:
                            if self.info_doc[i][0].get() > 0:

                                hoja[lis[j]+str(i+9)]=self.info_doc[i][j].get()

                            else:
                                hoja[lis[1]+str(i+9)]=self.info_doc[i][1].get()
                                j=4
                        
                        except:
                            if self.info_doc[i][1].get():

                                hoja[lis[1]+str(i+9)]=self.info_doc[i][1].get()
                                j=4
                       

                fecha=[]
                hoja["D3"]=self.Num_Fact.get()
                
                #cambia el formato de la fecha actual
                for i in self.Fecha_A.get():
                
                    if(i != "," and i != "(" and i != ")" and i != "'" and i !=" "):
                        fecha.append(i)
            
            
                hoja["D4"]=''.join(fecha)
                hoja["D6"]=self.VarTipo.get()
            
                #hoja["H5"]=self.Fecha_E.get()

                for i in range(3):

                    hoja[lis[3]+str(i+23)]=self.VarTotal[i].get()

                hoja["G1"]=self.Clien.get()
                hoja["G2"]=self.VarSede.get()
                hoja["G3"]=self.VarNit.get()
                hoja["G4"]=self.VarTlf.get()
                hoja["G5"]=self.VarCel.get()
                hoja["G6"]=self.Var_Direc.get()

                excel_doc.save(self.nombre_arch+" "+self.Num_Fact.get()+'.xlsx')


                messagebox.showinfo("Guardar Factura","Se ha guardado con exito la Factura")

        else:

            messagebox.showinfo("Error","No ha dado nombre a la Factura")


    #funcion: se encarga de verificar cuando se va llenando los datos int para ir actualizando
    def Actualizar(self,i=0, aux=0):
        
        if(i<self.tam):  
            #se encarga de verificar la exception cuando la variable queda vacia
            try:    #se actualiza el valor total
                self.info_doc[i][3].set(round(self.info_doc[i][0].get()*self.info_doc[i][2].get(),2))
                
                aux=aux+self.info_doc[i][3].get()
            
                self.aux=self.ventana.after(50,self.Actualizar,i+1, aux)

            except:
        
                self.aux=self.ventana.after(50,self.Actualizar,i+1, aux)

        else:
            try:
                aux2=aux-self.VarTotal[1].get()
                self.VarTotal[0].set(aux)
                self.VarTotal[2].set(aux2)
                self.ventana.after_cancel(self.aux)
                self.aux=self.ventana.after(50,self.Actualizar,i-i,aux-aux)

            except:
                #aux2=aux-self.VarTotal[1].get()
                self.VarTotal[0].set(aux)
                #self.VarTotal[2].set(aux2)
                self.ventana.after_cancel(self.aux)
                self.aux=self.ventana.after(50,self.Actualizar,i-i,aux-aux)

    def cargar_Num_fac(self):

        if(path.exists("fact.txt")==True):
            
            archivo=open("fact.txt","r")
            lista='*'
            while(lista!=''):
            
                lista=archivo.readline()
           
                if(lista != ''):
            
                    N=lista

            N_Factura=int(N)

            N_Factura+=1  
        
            archivo.close()

            aux2=str(N_Factura)
        
            archivo=open("fact.txt","a")
            archivo.write("\n")
            archivo.write(str(aux2))
            archivo.close()

            self.Num_Fact.set(aux2)
        
        else:

            messagebox.showinfo("Error","No existe un archivo para leer el numero de factura, cree uno en la opcion crear-archivo")


    def Crear_Archivo(self):

        valor=messagebox.askokcancel("Aceptar","¿Desea crear un nuevo archivo?")
        
        if valor==True:
            
            self.cargarYcerrar("10000")
        

    def cargarYcerrar(self,N):
        #self.Dir=input("ingrese donde se guardara el archivo: ")
        
        if(path.exists("fact.txt")==False):
            
            self.archivo=open("fact.txt","a")
            self.archivo.writelines(N)
            #self.archivo.write("\n")
            self.archivo.close()
            self.Num_Fact.set(N)
            messagebox.showinfo("Generar Archivo","Se ha creado el archo con exito")

        else:

            messagebox.showinfo("Error","Ya se encuentra un archivo creado con este nombre")

    def cargar_informacion(self):
        self.file=''
        self.file=filedialog.askopenfilename(title="selecciona un archivo",filetypes=())
        
        if(self.file!=''):
        
            print(self.file)
            lis=["A","B","G","H"]
            excel_doc=openpyxl.load_workbook(self.file)
            hoja=excel_doc.get_sheet_by_name('Hoja1')
        
            self.Num_Fact.set(hoja["D3"].value)
            self.Fecha_A.set(hoja["D4"].value)
            self.VarTipo.set(hoja["D6"].value)
            
            #self.Fecha_E.set(hoja["H5"].value)

            self.Clien.set(hoja["G1"].value)
            self.VarSede.set(hoja["G2"].value)
            self.VarNit.set(hoja["G3"].value)
            self.VarTlf.set(hoja["G4"].value)
            self.VarCel.set(hoja["G5"].value)
            self.Var_Direc.set(hoja["G6"].value)


            for i in range(self.tam):  #ciclo para iniciar varibles para almacenar 
            
                for j in range(3):
                
                    if((hoja["A"+str(i+9)]).value!=0 and (hoja["A"+str(i+9)]).value!=None):
                    
                        self.info_doc[i][j].set((hoja[lis[j]+str(i+9)]).value)

                    elif((hoja["B"+str(i+9)]).value!=None):

                        self.info_doc[i][1].set((hoja[lis[1]+str(i+9)]).value)


def Cambiar_Tam(ventana,aux):

    ventana.after_cancel(aux)
    ventana.destroy()
    vent_sec=Tk()
    tam=StringVar()
    vent_sec.geometry("200x100")

    entry=Entry(vent_sec,textvariable=tam)
    entry.place(x=50,y=40)
    label=Label(vent_sec,text="Ingrese el tam de la factura")
    label.place(x=40,y=20)

    boton=Button(vent_sec,text="Aceptar",command=lambda:iniciar(vent_sec,tam.get()))
    boton.place(x=50,y=60)

    vent_sec.mainloop()

def iniciar(vent_sec,tam):
    print(tam)
    vent_sec.destroy()
    factura=facturacion(int(tam))
  
def Nuevo(ventana, aux):
    ventana.after_cancel(aux)
    ventana.destroy()
    facturar=facturacion(10)  


facturar=facturacion(10)
#C:/Users/admin/Desktop/python/enero 2021/facturacion/fact.txt